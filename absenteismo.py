import tkinter as tk
from tkinter import filedialog, ttk
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
import openpyxl
import os
import datetime

# Função para abrir o diálogo de seleção de arquivo e registrar a seleção na text area
def select_file(entry, label):
    file_path = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if file_path:
        entry.delete(0, tk.END)
        entry.insert(0, file_path)
        
        message = f"{label} --> {file_path}\n"
        text_area.insert(tk.END, message)
        text_area.yview(tk.END)

# Função para selecionar o caminho para salvar: Arquivo ou Diretório
def select_save_path(entry, save_option):
    if save_option.get() == "Arquivo":
        path = filedialog.askopenfilename(
            filetypes=[("Arquivos Excel", "*.xlsx")],
            title="Selecione um arquivo Excel ou crie um novo"
        )
        
        if path:
            entry.delete(0, tk.END)
            entry.insert(0, path)
            message = f"Arquivo existente selecionado para salvar: {path}\n"
            text_area.insert(tk.END, message)
            text_area.yview(tk.END)
    
    elif save_option.get() == "Diretório":
        dir_path = filedialog.askdirectory(title="Selecione o diretório para salvar o arquivo")
        if dir_path:
            entry.delete(0, tk.END)
            output_file = os.path.join(dir_path, "absenteismo.xlsx")
            entry.insert(0, output_file)
            message = f"Novo arquivo será criado no diretório: {output_file}\n"
            text_area.insert(tk.END, message)
            text_area.yview(tk.END)

# Função para converter tempo no formato HH:MM para decimal
def time_to_decimal(time_str):
    try:
        if isinstance(time_str, str) and ':' in time_str:
            hours, minutes = map(int, time_str.split(':'))
            result = hours + minutes / 60
            return round(result, 2)
        elif isinstance(time_str, (int, float)):
            return round(float(time_str), 2)
        else:
            return 0
    except ValueError:
        return 0



# Função para ajustar a largura das colunas
def ajustar_largura_colunas(ws, df):
    for i, col in enumerate(df.columns, 1):
        max_length = max(df[col].astype(str).map(len).max(), len(col))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = adjusted_width

# Função para formatar células como porcentagem
def formatar_porcentagem(ws, colunas_percentual):
    for col in colunas_percentual:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = '0.00%' 

# Função para o botão "Executar"
def executar():
    # Obtém os caminhos dos arquivos das entradas
    listagem_simples = entry1.get()
    ausencias_justificadas = entry2.get()
    ausencias_injustificadas = entry3.get()
    save_path = entry_save.get()
    mes_selecionado = mes_combobox.get()
    ano_selecionado = ano_combobox.get()
    
    if not all([listagem_simples, ausencias_justificadas, ausencias_injustificadas, save_path, mes_selecionado]):
        text_area.insert(tk.END, "Erro: Todos os arquivos precisam ser selecionados e o mês deve ser escolhido.\n")
        text_area.yview(tk.END)
        return

    try:
        # 1. Carrega a planilha "listagemSimples"
        listagem_df = pd.read_excel(listagem_simples, header=3)

        # Filtra as colunas necessárias: Nº FOLHA, NOME, FUNÇÃO, DEPARTAMENTO
        listagem_df = listagem_df[['Nº FOLHA', 'NOME', 'FUNÇÃO', 'DEPARTAMENTO']]

        # 2. Remove a linha com o nome "TOTAL" na coluna "NOME"
        listagem_df = listagem_df[listagem_df['Nº FOLHA'] != 'TOTAL']

        # Caso tenha um espaço extra após a linha TOTAL, removemos linhas vazias ou extras
        # Verificamos se o DataFrame tem alguma linha vazia (caso haja uma linha em branco)
        listagem_df = listagem_df.dropna(how='all')
        
        # 2. Cria um novo DataFrame para a planilhaNova
        planilha_nova = listagem_df.copy()

        # Dicionário de tradução dos meses de português para inglês
        meses_traduzidos = {
            "JANEIRO": "January", "FEVEREIRO": "February", "MARÇO": "March", "ABRIL": "April", "MAIO": "May",
            "JUNHO": "June", "JULHO": "July", "AGOSTO": "August", "SETEMBRO": "September", "OUTUBRO": "October",
            "NOVEMBRO": "November", "DEZEMBRO": "December"
        }
        
        # 3. Cria a nova coluna "DATA" com o valor 01 do mês e ano selecionados
        mes_ingles = meses_traduzidos.get(mes_selecionado.upper(), "")  # Traduz o mês selecionado para inglês

        if mes_ingles:
            data_str = f"01-{mes_ingles}-{ano_selecionado}"  # Monta a string de data no formato correto
            data_obj = datetime.datetime.strptime(data_str, "%d-%B-%Y").date()  # Converte a string para um objeto datetime
        else:
            # Caso não encontre o mês, exibe um erro
            text_area.insert(tk.END, "Erro: Mês selecionado inválido.\n")
            text_area.yview(tk.END)
            return

        # Adiciona a nova coluna "DATA" na posição A (primeira coluna)
        planilha_nova.insert(0, 'DATA', data_obj)
        
        # 3. Carrega a planilha "ausências justificadas"
        ausencias_df = pd.read_excel(ausencias_justificadas, header=3)

        # Filtra as colunas necessárias: PREVISTAS, TRABALHADAS, AUSÊNCIAS, DESCRIÇÃO
        ausencias_df = ausencias_df[['PREVISTAS', 'TRABALHADAS', 'AUSÊNCIAS', 'DESCRIÇÃO']]

        # 4. Renomeia a coluna "AUSÊNCIAS" para "AUSÊNCIAS JUSTIFICADAS"
        ausencias_df.rename(columns={'AUSÊNCIAS': 'AUSÊNCIAS JUSTIFICADAS'}, inplace=True)

        # 5. Converte as colunas de tempo para decimal
        for col in ['PREVISTAS', 'TRABALHADAS', 'AUSÊNCIAS JUSTIFICADAS']:
            ausencias_df[col] = ausencias_df[col].apply(time_to_decimal)

        # 6. Relaciona a coluna "DESCRIÇÃO" com "NOME"
        planilha_nova = planilha_nova.merge(ausencias_df, left_on='NOME', right_on='DESCRIÇÃO', how='left')

        # 7. Remove a coluna "DESCRIÇÃO"
        planilha_nova = planilha_nova.drop(columns=['DESCRIÇÃO'])

        # 8. Carrega a planilha "ausências injustificadas"
        ausencias_injustificadas_df = pd.read_excel(ausencias_injustificadas, header=3)

        # Filtra as colunas necessárias: AUSÊNCIAS, DESCRIÇÃO
        ausencias_injustificadas_df = ausencias_injustificadas_df[['AUSÊNCIAS', 'DESCRIÇÃO']]

        # 9. Renomeia a coluna "AUSÊNCIAS" para "AUSÊNCIAS INJUSTIFICADAS"
        ausencias_injustificadas_df.rename(columns={'AUSÊNCIAS': 'AUSÊNCIAS INJUSTIFICADAS'}, inplace=True)

        # 10. Converte para decimal
        ausencias_injustificadas_df['AUSÊNCIAS INJUSTIFICADAS'] = ausencias_injustificadas_df['AUSÊNCIAS INJUSTIFICADAS'].apply(time_to_decimal)

        # 11. Relaciona a coluna "DESCRIÇÃO"
        planilha_nova = planilha_nova.merge(ausencias_injustificadas_df, left_on='NOME', right_on='DESCRIÇÃO', how='left')

        # 12. Remove a coluna "DESCRIÇÃO"
        planilha_nova = planilha_nova.drop(columns=['DESCRIÇÃO'])

                # 13. Cálculos de absenteísmo com verificação para PREVISTAS == 0
        planilha_nova['ABSENTEÍSMO GERAL'] = planilha_nova.apply(
            lambda row: 0 if row['PREVISTAS'] == 0 else (row['AUSÊNCIAS JUSTIFICADAS'] + row['AUSÊNCIAS INJUSTIFICADAS']) / row['PREVISTAS'], axis=1
        )
        planilha_nova['ABSENTEÍSMO JUSTIFICADO'] = planilha_nova.apply(
            lambda row: 0 if row['PREVISTAS'] == 0 else row['AUSÊNCIAS JUSTIFICADAS'] / row['PREVISTAS'], axis=1
        )
        planilha_nova['ABSENTEÍSMO INJUSTIFICADO'] = planilha_nova.apply(
            lambda row: 0 if row['PREVISTAS'] == 0 else row['AUSÊNCIAS INJUSTIFICADAS'] / row['PREVISTAS'], axis=1
        )

        # 14. Verifica se o arquivo já existe
        output_file = save_path

        if os.path.exists(output_file):
            # Se o arquivo já existe, abre a planilha existente
            wb = load_workbook(output_file)

            # Verifica se a aba já existe
            nome_aba = f"{mes_selecionado}-{ano_selecionado}"
            if nome_aba in wb.sheetnames:
                text_area.insert(tk.END, f"Erro: A aba '{nome_aba}' já existe na planilha.\n")
                text_area.yview(tk.END)
                return

            # Cria uma nova aba com o nome do mês
            ws = wb.create_sheet(title=nome_aba)

        else:
            # Se o arquivo não existe, cria um novo arquivo
            nome_aba = f"{mes_selecionado}-{ano_selecionado}"
            wb = Workbook()
            ws = wb.active
            ws.title = nome_aba

        # 15. Preenche os dados na nova aba
        for r in dataframe_to_rows(planilha_nova, index=False, header=True):
            ws.append(r)

        # 16. Ajusta a largura das colunas e aplica o cabeçalho em negrito
        ajustar_largura_colunas(ws, planilha_nova)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 17. Formata a coluna "DATA" como data no formato "JAN/2024"
        coluna_data = planilha_nova.columns.get_loc('DATA') + 1  # Converte o nome da coluna "DATA" para o índice
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=coluna_data)
            cell.number_format = '[$-pt-BR]mmm/yyyy'
        

        # 17. Formata as colunas de percentual
        colunas_percentual = [planilha_nova.columns.get_loc('ABSENTEÍSMO GERAL') + 1,
                              planilha_nova.columns.get_loc('ABSENTEÍSMO JUSTIFICADO') + 1,
                              planilha_nova.columns.get_loc('ABSENTEÍSMO INJUSTIFICADO') + 1]
        formatar_porcentagem(ws, colunas_percentual)

        # 18. Salva o arquivo
        wb.save(output_file)

        text_area.insert(tk.END, f"Processamento concluído. Arquivo salvo em: {output_file}\n")
        text_area.yview(tk.END)

    except Exception as e:
        text_area.insert(tk.END, f"Erro ao processar os arquivos: {str(e)}\n")
        text_area.yview(tk.END)

# Configuração da janela principal
root = tk.Tk()
root.title("ABSENTEÍSMO")
root.geometry("450x650")
root.resizable(False, False)

main_frame = ttk.Frame(root, padding="20")
main_frame.pack(fill="both", expand=True)

# Seleção do logotipo
try:
    logo_img = tk.PhotoImage(file="logo.png")
    logo_label = tk.Label(main_frame, image=logo_img)
    logo_label.pack(pady=(0, 0))
except Exception as e:
    print("Erro ao carregar a imagem:", e)

# Campo para a seleção da planilha listagem simples
label1 = ttk.Label(main_frame, text="Selecione a planilha listagem simples:")
label1.pack(anchor="w", pady=(0, 5))
frame1 = ttk.Frame(main_frame)
frame1.pack(fill="x", pady=(0, 10))
entry1 = ttk.Entry(frame1)
entry1.pack(side="left", fill="x", expand=True)
button1 = ttk.Button(frame1, text="...", command=lambda: select_file(entry1, "Listagem simples"), width=3)
button1.pack(side="right")

# Campo para a seleção da planilha ausências justificadas
label2 = ttk.Label(main_frame, text="Selecione a planilha ausências justificadas:")
label2.pack(anchor="w", pady=(0, 5))
frame2 = ttk.Frame(main_frame)
frame2.pack(fill="x", pady=(0, 10))
entry2 = ttk.Entry(frame2)
entry2.pack(side="left", fill="x", expand=True)
button2 = ttk.Button(frame2, text="...", command=lambda: select_file(entry2, "Ausências justificadas"), width=3)
button2.pack(side="right")

# Campo para a seleção da planilha ausências injustificadas
label3 = ttk.Label(main_frame, text="Selecione a planilha ausências injustificadas:")
label3.pack(anchor="w", pady=(0, 5))
frame3 = ttk.Frame(main_frame)
frame3.pack(fill="x", pady=(0, 10))
entry3 = ttk.Entry(frame3)
entry3.pack(side="left", fill="x", expand=True)
button3 = ttk.Button(frame3, text="...", command=lambda: select_file(entry3, "Ausências injustificadas"), width=3)
button3.pack(side="right")

# Frame para os radio buttons
frame_radiobuttons = ttk.Frame(main_frame)
frame_radiobuttons.pack(anchor="w", pady=(5, 10))

# Radio buttons para escolha de arquivo ou diretório
save_option = tk.StringVar(value="Arquivo")

radio_file = ttk.Radiobutton(frame_radiobuttons, text="Arquivo existente", variable=save_option, value="Arquivo")
radio_file.pack(side="left", padx=10)

radio_dir = ttk.Radiobutton(frame_radiobuttons, text="Novo arquivo", variable=save_option, value="Diretório")
radio_dir.pack(side="left", padx=10)

# Campo para a seleção do diretório de salvamento
label_save = ttk.Label(main_frame, text="Selecione onde salvar o arquivo:")
label_save.pack(anchor="w", pady=(0, 5))
frame_save = ttk.Frame(main_frame)
frame_save.pack(fill="x", pady=(0, 25))
entry_save = ttk.Entry(frame_save)
entry_save.pack(side="left", fill="x", expand=True)

# Botão para selecionar caminho de salvamento
button_save = ttk.Button(frame_save, text="...", command=lambda: select_save_path(entry_save, save_option), width=3)
button_save.pack(side="right")

# Frame para agrupar "Mês" e "Ano" na mesma linha
frame_mes_ano = ttk.Frame(main_frame)
frame_mes_ano.pack(fill="x", pady=(10, 20))

# Label para "Mês"
label_mes = ttk.Label(frame_mes_ano, text="Mês:")
label_mes.pack(side="left", padx=(0, 10))

# Combobox para selecionar o mês
mes_combobox = ttk.Combobox(frame_mes_ano, values=["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"])
mes_combobox.pack(side="left", fill="x", expand=True, padx=(0, 10))

# Label para "Ano"
label_ano = ttk.Label(frame_mes_ano, text="Ano:")
label_ano.pack(side="left", padx=(10, 5))

# Combobox para selecionar o ano
ano_combobox = ttk.Combobox(frame_mes_ano, values=["2024", "2025", "2026", "2027", "2028", "2029", "2030"])
ano_combobox.pack(side="left", padx=(5, 0))

# Ajustando o layout geral
frame_mes_ano.pack(fill="x", pady=(0, 10))

# Área de texto para log de saída
label4 = ttk.Label(main_frame, text="Log de saída:")
label4.pack(anchor="w", pady=(10, 5))
text_area = tk.Text(main_frame, height=4, wrap="word", font=("Arial", 8))
text_area.pack(fill="x", pady=(0, 20))

# Botão para executar o processo
executar_button = tk.Button(
    main_frame, 
    text="Executar", 
    command=executar, 
    bg="#014184", 
    fg="white", 
    font=("Arial", 10, "bold"), 
    padx=10, 
    pady=10,
    relief="flat"
)
executar_button.config(width=10)
executar_button.pack(pady=(0, 0))

root.mainloop()